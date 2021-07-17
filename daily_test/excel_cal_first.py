#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/3/9 17:36
# @Author  : xin
# @Detail  : 读取Excel表并进行计算后输出Excel报告文档

import os  # 用于获取Excel文件路径，以及对分析后的Excel文件进行保存
import openpyxl  # 可以对Excel进行操作
import xlsxwriter
import xlrd
import datetime  # 日期
import time  # 时间

sheet_list = ['旧否新否', '旧否新是', '旧是新否', '旧是新是对比成功', '旧是新是对比失败']
file_path = r"/Users/huangxiaoxin/Desktop/files/01_cass"
file_name = "奔驰sbom对比新旧译码数据对比.xlsx"
file_result = os.path.join(file_path, file_name)


# 创建新的sheet存放筛选过的数据
def creat_new_sheet():
    # 获取Excel文件生成对象obj_excel
    obj_excel = openpyxl.load_workbook(file_result)
    # 获取当前工作簿生成对象obj_sheet
    obj_sheet = obj_excel.active
    # 设置参数，用以循环生成新sheet
    index = 0

    header_list = []
    for c in range(1,obj_sheet.max_column+1):
        v = obj_sheet.cell(1,c).value
        header_list.append(v)
    print(header_list)

    # 以sheet_list长度，筛选条件循环生成新sheet
    for i in range(len(sheet_list)):
        count = 2
        sh1 = obj_excel.create_sheet(str(sheet_list[index]))

        # 将原表头写入新表
        for num in range(len(header_list)):
            sh1.append(header_list)

        for rows in obj_sheet.rows:
            if rows[3].coordinate != 'D1' and rows[3].value == '是':
                sh1['A' + str(count)] = rows[0].value
                sh1['B' + str(count)] = rows[1].value
                sh1['C' + str(count)] = rows[2].value
                sh1['D' + str(count)] = rows[3].value
                sh1['E' + str(count)] = rows[4].value
                sh1['F' + str(count)] = rows[5].value
                sh1['G' + str(count)] = rows[6].value
                sh1['H' + str(count)] = rows[7].value
                sh1['I' + str(count)] = rows[8].value
                sh1['J' + str(count)] = rows[9].value
                # sh1['K' + str(count)] = rows[10].value
                # sh1['L' + str(count)] = rows[11].value

                print(f'正在分析{sheet_list[index]}表的数据......')
                count += 1
        index += 1
    obj_excel.save('/Users/huangxiaoxin/Desktop/files/01_cass/奔驰sbom对比新旧译码数据对比.xlsx')


if __name__ == "__main__":
    start_time = time.time()
    creat_new_sheet()
    print(f'分析完成，用时为{time.time() - start_time}秒')